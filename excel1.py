from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
#建立新的excel 檔
#wb= Workbook()
wb=load_workbook("2.xlsx")
ws=wb.active
ws.title="No"

#ws.append([1,2,34,56,789]) #新增橫排的資料
#ws.append([1,2,34,526,789])
#ws.append([1,2,34,256,789])
#ws.append([1,2,34,526,789])

for row in range(1,5):
    for col in range(1,6):
        char=  get_column_letter(col)
        print(ws[char + str(row)].value)

ws.merge_cells("A1:E1") #合併儲存格
ws.unmerge_cells("A1:E1") #復原

#ws.insert_rows(3) #插入橫排
#ws.delete_rows(3) #刪除橫排

ws.move_range("A3:E4", rows=2,cols=3)

wb.save("2.xlsx")
